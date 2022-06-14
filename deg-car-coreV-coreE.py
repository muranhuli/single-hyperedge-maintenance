from openpyxl import Workbook
import math
import os


# file = 'S_CoGe  S_CoHi  S_DAWN  S_DrAb  S_MaAn  S_MiAc  S_NDC  S_StAn  S_TrCl  S_WaTr T_CoBi T_CoDB T_ThAU T_ThSO'.split()
file ='MaAn ThAU ThMS CoMH'.split()
filepath = '../graphData/hypergraph/'
wb = Workbook()
wb.create_sheet('deg')
wb.create_sheet('car')
wb.create_sheet('core')
wb.create_sheet('coreE')
count =0
for name in file:
    count +=1
    filename = filepath + name
    hyperedge = {}
    index = 0
    hypernode = {}
    with open(filename) as f:
        for line in f:
            if len(line.split()) ==1:
                continue
            hyperedge[index]=line.split()
            for node in hyperedge[index]:
                if node not in hypernode:
                    hypernode[node]=[]
                hypernode[node].append(index)
            index += 1
    deg = {}
    deg_c =[]
    maxDeg=0
    for info in hypernode.values():
        if len(info) not in deg:
            deg[len(info)]=0
        deg[len(info)]+=1
        maxDeg=max(maxDeg,len(info))
    for i in range(maxDeg+1):
        deg_c.append(0)
    for key,value in deg.items():
        deg_c[key]=value
    # for i in range(1,maxDeg+1):
    #     deg_c[i]+=deg_c[i-1]
    # for i in range(maxDeg+1):
    #     deg_c[i]=deg_c[i]/deg_c[maxDeg]   
    
    car ={}
    for info in hyperedge.values():       
        if len(info) == 1:
            print(info)
            exit(-1)
        if len(info) not in car:
            car[len(info)]=0
        car[len(info)]+=1
        deg = {}
    car_c =[]
    maxCar = 0
    for info in hyperedge.values():
        if len(info) not in car:
            car[len(info)]=0
        car[len(info)]+=1
        maxCar=max(maxCar,len(info))
    for i in range(maxCar+1):
        car_c.append(0)
    for key,value in car.items():
        car_c[key]=value
    # for i in range(1,maxCar+1):
    #     car_c[i]+=car_c[i-1]
    # for i in range(maxCar+1):
    #     car_c[i]=car_c[i]/car_c[maxCar]
    
    core ={}
    maxCore = 0
    with open(f'./result/CoreNumber/{name}-coreVertexValue') as f:
        for line in f:
            a,b=line.split()
            if int(b) not in core:
                core[int(b)]=0
            core[int(b)]+=1
            maxCore=max(maxCore,int(b))
    core_c =[]
    for i in range(maxCore+1):
        core_c.append(0)
    for key,value in core.items():
        core_c[key]=value
    for i in range(1,maxCore+1):
        core_c[i]+=core_c[i-1]
    for i in range(maxCore+1):
        core_c[i]=core_c[i]/core_c[maxCore]
        
    coreE ={}
    maxCoreE = 0
    with open(f'./result/CoreNumber/{name}-coreEdgeValue') as f:
        for line in f:
            b=line.split()[0]
            if int(b) not in coreE:
                coreE[int(b)]=0
            coreE[int(b)]+=1
            maxCoreE=max(maxCoreE,int(b))
    coreE_c =[]
    for i in range(maxCoreE+1):
        coreE_c.append(0)
    for key,value in coreE.items():
        coreE_c[key]=value
    for i in range(1,maxCoreE+1):
        coreE_c[i]+=coreE_c[i-1]
    for i in range(maxCore+1):
        coreE_c[i]=coreE_c[i]/coreE_c[maxCoreE]
        
        
        

    deg_c.pop(0)
    ws=wb['deg']
    for index, value in enumerate(deg_c, 1):
        ws.cell(index,count).value=value
        
    car_c.pop(0)
    ws=wb['car']
    for key,value in enumerate(car_c, 1):
        ws.cell(key,count).value=value
    
    core_c.pop(0)
    ws=wb['core']
    for key,value in enumerate(core_c, 1):
        ws.cell(key,count).value=value*100
        
    coreE_c.pop(0)
    ws=wb['coreE']
    for key,value in enumerate(coreE_c, 1):
        ws.cell(key,count).value=value*100
    print(name)
ws = wb["Sheet"]
wb.remove(ws)
wb.save(f'./result/deg-car-coreV-coreE.xlsx')